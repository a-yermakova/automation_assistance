"""Module of scripts helping with process of automation an analyst's model."""

import re
import openpyxl
import configparser
from io import BytesIO
from pathlib import Path
from typing import Generator
from dataclasses import dataclass

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from automation_assistance_exceptions import EmptyTagCellInModel


def is_empty_array(array: [list[Cell]|tuple[Cell]]) -> bool:
    """Проверяет, является ли ряд или столбец пустым (значения ячеек None)."""
    checklist = [cell.value for cell in array if cell.value is not None]
    return len(checklist) == 0


def define_first_prognosis_column_number(worksheet, explicit_row=None, start_column: int = 1) -> int:
    """Определяем первый найденный при обходе слева-направо столбец с прогнозами, у которого в хэдэре
     есть обозначение прогнозного периода (F или П)."""
    date_row = explicit_row or worksheet.min_row
    # Используем свойство функции iter_rows по обходу ряда слева-направо. min_col - для поиска только по датам
    for row in worksheet.iter_rows(min_row=date_row, max_row=date_row, min_col=start_column):
        for cell in row:
            # 'П' добавлена для российских компаний
            if isinstance(cell.value, str) and (cell.value.endswith('F') or cell.value.endswith('П')):
                return cell.column
    raise ValueError(f'На странице "{worksheet.title}" не найден период с отметкой F или П')


@dataclass
class Equivalent:
    """Класс для хранения информации о названиях статьи из разных источников."""

    # Название статьи из переданного отчета эмитента
    tag_from_filling: str
    # Название статьи в переданной аналитической модели
    model_tag: str
    # Список названий статей из левой части конфига конкретного переданного
    # источника данных (XBRL, PDF, XLSX), где в значениях (правая часть конфига)
    # встретилось название статьи из переданного отчета эмитента
    tags_from_config: list = None


def find_first_numeric_cell_in_column(column_for_searching: Generator) -> str | None:
    """ Возвращает адрес первой ячейки в столбце, имеющей числовое значение
    Args:
        column_for_searching: итератор по ячейкам в колонке
    Returns:
        координаты первой числовой ячейки
    """
    for cell in column_for_searching:
        if (isinstance(cell.value, int) or isinstance(cell.value, float)) and cell.value != 0:
            coordinates_of_first_numeric_cell = cell.coordinate
            return coordinates_of_first_numeric_cell


def find_coordinates_for_start_in_filling(worksheet: Worksheet, cell_value_to_find: int|float) -> str | None:
    """Производит поиск ячейки с определенным значением и возвращает координаты первой числовой ячейки в этом столбце
    Args:
        :param worksheet: страница XLSX файла для поиска
        :type worksheet:
        :param cell_value_to_find: значение ячейки, которую необходимо найти
        :type cell_value_to_find: int
    Returns:
        координаты найденной ячейки
    """
    # min_col=2, т.к. первый столбец вероятнее всего столбец со статьями (str), поэтому пропускаем
    for column in worksheet.iter_cols(min_col=2):
        for general_cell in column:
            if general_cell.value == cell_value_to_find:
                cell_in_filling_column_index = general_cell.column
                # получить номер столбца, в котором найдено совпадающее значение
                # найти в этом столбце координату верхней числовой ячейки
                for cell in next(worksheet.iter_cols(min_col=cell_in_filling_column_index,
                                                     max_col=cell_in_filling_column_index)):
                    if isinstance(cell.value, int) or isinstance(cell.value, float):
                        coordinates_of_cell_for_start_in_filling = cell.coordinate
                        return coordinates_of_cell_for_start_in_filling


def get_coordinates_of_cells_in_column_with_same_fiscal_period(selected_sheet_in_model_wb: Worksheet,
                                                               selected_sheet_in_issuer_wb: Worksheet,
                                                               number_of_column_including_needed_cell: int
                                                               ) -> tuple[str, str]:
    """Итерируется по ячейкам определенного столбца в файле аналитической модели до тех пор, пока не найдет
     первую непустую ячейку с числовым значением. Далее вызывает функцию для поиска ячейки с таким же значением
     на выбранном листе файла отчёта эмитента. Сохраняет координаты обеих ячеек. В случае, если координаты
     не найдены, рассчитывается значение по умолчанию.

    Args:
        selected_sheet_in_model_wb: страница XLSX файла аналитической модели для поиска
        selected_sheet_in_issuer_wb: страница XLSX файла отчёта эмитента для поиска
        number_of_column_including_needed_cell: номер колонки для поиска ячейки

    Returns:
        координаты ячеек для старта в аналитической модели и в отчете эмитента
    """
    # находим координаты первой ячейки с численным значением (ячейка для начала обработки)
    for cell in next(selected_sheet_in_model_wb.iter_cols(min_col=number_of_column_including_needed_cell,
                                                          max_col=number_of_column_including_needed_cell)):
        if (isinstance(cell.value, int) or isinstance(cell.value, float)) and cell.value != 0:
            coordinates_of_cell_for_start_in_model = cell.coordinate
            coordinates_of_cell_for_start_in_filling = find_coordinates_for_start_in_filling(
                                                                        selected_sheet_in_issuer_wb, cell.value)
            if coordinates_of_cell_for_start_in_filling:
                return coordinates_of_cell_for_start_in_model, coordinates_of_cell_for_start_in_filling
    # Если координаты не вычислены, передается значение по умолчанию
    else:
        # для заполнения по умолчанию берем столбец B (то есть второй)
        index_of_column_for_autofill = 2
        column_for_autofill_in_filling = next(selected_sheet_in_issuer_wb.iter_cols(
                                                          min_col=index_of_column_for_autofill,
                                                          max_col=index_of_column_for_autofill))
        # ищем верхнюю ячейку с числовым значением
        default_address_in_filling = find_first_numeric_cell_in_column(column_for_autofill_in_filling)
        # странным сложным способом получаем букву нужной колонки
        default_column_letter = next(selected_sheet_in_model_wb.iter_cols(min_col=number_of_column_including_needed_cell,
                                                                          max_col=number_of_column_including_needed_cell)
                                                                          )[0].column_letter
        return f'{default_column_letter}4', default_address_in_filling


def specify_function_for_cell_address_searching(model_binary_stream: BytesIO, issuer_binary_stream: BytesIO,
                                                selected_model_sheet: str, selected_issuer_sheet: str):
    """ В зависимости от переданного названия листа производит поиск координат ячейки для начала обработки
    Args:
        model_binary_stream: бинарный поток XLSX файла аналитической модели
        issuer_binary_stream: бинарный поток XLSX файла отчета эмитента
        selected_model_sheet: выбранный пользователем лист в файле аналитической модели
        selected_issuer_sheet: выбранный пользователем лист в файле отчета эмитента

    Returns:
        координаты ячеек для начала обработки в файле аналитической модели и в файле отчёта эмитента
    """
    model_workbook = openpyxl.load_workbook(model_binary_stream)
    issuer_workbook = openpyxl.load_workbook(issuer_binary_stream)
    selected_sheet_in_model_wb = model_workbook[selected_model_sheet]
    selected_sheet_in_issuer_wb = issuer_workbook[selected_issuer_sheet]
    if 'Баланс' in selected_sheet_in_model_wb.title or 'Balance' in selected_sheet_in_model_wb.title:
        index_of_last_column = selected_sheet_in_model_wb.max_column
        # проверяем последнюю непустую колонку действительно ли она непустая
        # получаем номер колонки, которую необходимо обработать
        for column_index in range(index_of_last_column, 1, -1):
            for column in selected_sheet_in_model_wb.iter_cols(min_col=column_index, max_col=column_index):
                if not is_empty_array(column):
                    # если координаты не найдены, возвращается значение по умолчанию
                    address_start_cell_in_model, address_start_cell_in_filling = \
                        get_coordinates_of_cells_in_column_with_same_fiscal_period(selected_sheet_in_model_wb,
                                                                                   selected_sheet_in_issuer_wb,
                                                                                   column_index)
                    return address_start_cell_in_model, address_start_cell_in_filling
    else:
        # названия столбцов записаны во втором ряду
        row_in_model_to_start = 2
        first_prognosis_column_number = define_first_prognosis_column_number(selected_sheet_in_model_wb,
                                                                             row_in_model_to_start)
        # Берём прогнозный столбик - 1, т. к. нужен последний столбик с данными из отчетов
        # если координаты не найдены, возвращается значение по умолчанию
        address_start_cell_in_model, address_start_cell_in_filling = \
            get_coordinates_of_cells_in_column_with_same_fiscal_period(selected_sheet_in_model_wb,
                                                                       selected_sheet_in_issuer_wb,
                                                                       first_prognosis_column_number - 1)
        return address_start_cell_in_model, address_start_cell_in_filling


def add_similar_statement_tags_from_config(statement_block: str, data_source: str, list_of_equivalents: list):
    """По полученным аргументам выбирает какой файл config и источник
    данных необходимо использовать. Сравнивает переданные в списке list_of_equivalents
    названия статей из отчета эмитента с названиями статей из отчетов эмитентов из файла config.
    При наличии совпадений заносит список названий статей для аналитической модели в список tags_from_config
     в dataclass Equivalent.
    Args:
        :param statement_block: блок статей, который определяет, какой конфиг необходимо использовать
                                (Баланс, Финансовые результаты, Сегменты, Отчет о движении денежных средств)
        :type statement_block: str
        :param data_source: источник данных, который определяет раздел конфига (XBRL, XLSX, PDF)
        :type data_source: str
        :param list_of_equivalents: список, состоящий из объектов dataclass Equivalent
        :type list_of_equivalents: list
    """
    # создаём путь до общей папки с конфигами
    # data_folder = Path(__file__).parent.parent.resolve().joinpath('data')
    match statement_block:
        case 'Баланс':
            config_name = 'balance_config.ini'
            # path_to_config = data_folder.joinpath(config_name)
            path_to_config = config_name
        case 'Финансовые результаты':
            config_name = 'income_config.ini'
            # path_to_config = data_folder.joinpath(config_name)
            path_to_config = config_name
        case 'Сегменты':
            config_name = 'segments_config.ini'
            # path_to_config = data_folder.joinpath(config_name)
            path_to_config = config_name
        case 'Отчет о движении денежных средств':
            config_name = 'cashflow_config.ini'
            # path_to_config = data_folder.joinpath(config_name)
            path_to_config = config_name

    match data_source:
        case 'XBRL':
            used_data_source = 'XBRL template'
        case 'XLSX':
            used_data_source = 'XLSX statements'
        case 'PDF':
            used_data_source = 'PDF statements'

    used_config = configparser.ConfigParser()
    used_config.read(path_to_config, encoding='UTF-8')
    # Берем один dataclass из списка, из него берем тэг из отчета и сравниваем его со
    # всеми тэгами эмитентов из конфига (справа после равно), перебирая поэлементно словарь ({ключ: тэг, тэг, тэг}).
    for equivalent in list_of_equivalents:
        # Так как для каждого эквивалента нужно пройтись по всему конфигу и собрать все совпадения
        list_of_config_model_tags = []
        for config_model_tag, config_issuer_tags in used_config[used_data_source].items():
            # Если тэг из переданного отчета эмитента является частью одного из тэгов из правой части конфига,
            # добавляем в список list_of_config_model_tags соответствующее название статьи из левой части конфига
            if any(equivalent.tag_from_filling.lower() in value.lower() for value in config_issuer_tags.split('\n')):
                # записываем статью из конфига с заглавной буквы (чтоб красиво было :) )
                list_of_config_model_tags.append(config_model_tag.capitalize())
        # Полученный список присваиваем tags_from_config из dataclass.
        equivalent.tags_from_config = list_of_config_model_tags


def get_sheetnames_with_binary_stream(xlsx_binary_stream: BytesIO) -> list[str]:
    """Отдает список названий страниц из бинарного потока XLSX файла.
    Args:
        :param xlsx_binary_stream: бинарный поток XLSX файла,
        :type xlsx_binary_stream: BytesIO

    Returns:
        список названий страниц WB
    """
    workbook = openpyxl.load_workbook(xlsx_binary_stream)
    return workbook.sheetnames


def check_cell_address_input(address_of_start: str) -> [str or None]:
    """Функция для проверки введенного адреса верхней ячейки столбца числовых значений.
    Args:
        :param address_of_start: введенный адрес верхней ячейки столбца числовых значений,
        :type address_of_start: str
    Returns:
        стандартизированный адрес ячейки (тип строка) или None, если проверка не пройдена"""
    address_of_start = address_of_start.upper()
    address_of_start = address_of_start.strip()
    if not re.match(r'[A-Z]{1,3}\d{1,3}', address_of_start):
        return
    else:
        return address_of_start


def tags_equations_creator(*, model_binary_stream: BytesIO, issuer_binary_stream: BytesIO,
                           selected_model_sheet: str, selected_issuer_sheet: str,
                           model_address_of_start: str, issuer_address_of_start: str,
                           index_of_model_column_with_tags: int = 2,
                           index_of_issuer_column_with_tags: int = 1) -> list[tuple]:
    """Отдает список эквивалентных названий статей, совпадающих по значениям за одинаковый фискальный период
    Args:
        :param model_binary_stream: бинарный поток XLSX файла аналитической модели
        :type model_binary_stream: BytesIO
        :param issuer_binary_stream: бинарный поток XLSX файла отчета эмитента
        :type issuer_binary_stream: BytesIO
        :param selected_model_sheet: название выбранной для сравнения ws из файла аналитической модели
        :type selected_model_sheet: str
        :param selected_issuer_sheet: название выбранной для сравнения ws из файла отчета эмитента
        :type selected_issuer_sheet: str
        :param model_address_of_start: адрес верхней ячейки столбца для сравнения (со значениями статей за определенный
                                        фискальный период) в файле аналитической модели
        :type model_address_of_start: str
        :param issuer_address_of_start: адрес верхней ячейки столбца для сравнения (со значениями статей
                                         за определенный фискальный период) в файле отчета эмитента
        :type issuer_address_of_start: str
        :param index_of_model_column_with_tags: индекс столбца с названиями статей в аналитической модели
                                                 (по умолчанию 2)
        :type index_of_model_column_with_tags: int
        :param index_of_issuer_column_with_tags: индекс столбца с названиями статей в отчете эмитента (по умолчанию 1)
        :type index_of_issuer_column_with_tags: int
    Returns:
        список кортежей эквивалентных названий статей, где 1-ый элемент кортежа — название в аналитической модели, а
        2-ой элемент — название статьи в отчете эмитента
        """
    # создаем workbook из каждого бинарного потока
    model_workbook = openpyxl.load_workbook(model_binary_stream, data_only=True)
    issuer_workbook = openpyxl.load_workbook(issuer_binary_stream, data_only=True)
    # создаем worksheet по названию каждого из переданных листов
    model_worksheet = model_workbook[selected_model_sheet]
    issuer_worksheet = issuer_workbook[selected_issuer_sheet]
    # сохраняем номера столбцов и рядов для обхода столбца для сравнения
    model_index_of_column = model_worksheet[model_address_of_start].column
    model_index_of_row = model_worksheet[model_address_of_start].row
    issuer_index_of_column = issuer_worksheet[issuer_address_of_start].column
    issuer_index_of_row = issuer_worksheet[issuer_address_of_start].row
    # создаем два итератора для перебора ячеек одного столбца
    model_column = next(model_worksheet.iter_cols(min_col=model_index_of_column,
                                                  max_col=model_index_of_column,
                                                  min_row=model_index_of_row))
    issuer_column = next(issuer_worksheet.iter_cols(min_col=issuer_index_of_column,
                                                    max_col=issuer_index_of_column,
                                                    min_row=issuer_index_of_row))
    list_of_equivalents = []
    # сравниваем значение каждой ячейки в столбце из файла аналитической модели с ячейками из файла отчета эмитента
    for cell in model_column:
        cell_value = cell.value
        if cell_value is None:
            continue
        for comparative_cell in issuer_column:
            comparative_cell_value = comparative_cell.value
            if comparative_cell_value is None:
                continue
            # в ячейке может быть float - приведем все к целым числам (int)
            if isinstance(cell_value, float):
                cell_value = int(cell_value)
            if isinstance(comparative_cell_value, float):
                comparative_cell_value = int(comparative_cell_value)
            # если значения ячеек равны, приравниваем значения названий статей и добавляем в список строк(str)
            if cell_value == comparative_cell_value:
                model_cell = model_worksheet.cell(row=cell.row, column=index_of_model_column_with_tags)
                tag_cell = model_cell.value
                # чтобы не было случая, когда напротив найденного значения в столбце с тегами ничего не написано
                if not tag_cell:
                    raise EmptyTagCellInModel(f'Совпадение значений найдено, '
                                              f'но ячейка с названием статьи в модели ({model_cell.coordinate}) - пуста')
                # пропускаем процентные значения, т. к. они округляются до нуля
                if '%' in tag_cell:
                    continue
                # если ячейка уже залита жёлтым, т.е. прошла автоматизацию и закачивает значение - можно её уже не брать
                if model_cell.fill.fgColor.index == 'FFFFFF00':
                    continue
                tag_comparative_cell = issuer_worksheet.cell(row=comparative_cell.row,
                                                             column=index_of_issuer_column_with_tags).value

                list_of_equivalents.append(Equivalent(model_tag=str(tag_cell),
                                                      tag_from_filling=str(tag_comparative_cell)))
    return list_of_equivalents


if __name__ == '__main__':
    # xlsx = openpyxl.load_workbook(r'C:\Users\trainee_02\Desktop\work\FEES.xlsx')
    wb_model = open(r'C:\Users\trainee_02\Desktop\work\FEES.xlsx', 'rb')
    wb_issuer = open(r'C:\Users\trainee_02\Desktop\work\FEES_parsed.xlsx', 'rb')
    print(specify_function_for_cell_address_searching(wb_model, wb_issuer, 'FEES Модель', 'page-4-table-1'))


